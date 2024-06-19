'''б) Отсортируйте полученную матрицу в порядке убывания.'''

import openpyxl

new_xls = openpyxl.workbook.Workbook()
sheet1  = new_xls.create_sheet('Лист 1')
sheet2  = new_xls.create_sheet('Лист 2')

sheet1['A1'] = 1111
sheet1['A2'] = 2222
sheet1['A3'] = 3333

a = sheet1['A1'].value,sheet1['A2'].value ,sheet1['A3'].value
a = sorted(a,reverse=True)


sheet2['A1'] = a[0]
sheet2['A2'] = a[1]
sheet2['A3'] = a[2]
font = openpyxl.styles.Font(name='Tahoma', size=16, bold=True,
                italic=False, vertAlign=None, underline='none',
                                strike=False, color='FF0000FF')
sheet2['A1'].font = font
sheet2['A2'].font = font
sheet2['A3'].font = font



new_xls.save('MyExcel.xlsx')

